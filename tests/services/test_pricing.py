from copy import deepcopy
from tempfile import NamedTemporaryFile
from unittest import TestCase
from unittest.mock import ANY, patch

import openpyxl
import pytest
import responses
from connect.client import ClientError
from connect.client.rql import R
from openpyxl import load_workbook

from connect_ext_ppr.services.cbc_hub import CBCService
from connect_ext_ppr.services.pricing import (
    _determine_dataset,
    _fetch_and_validate_batch,
    _get_reseller_id,
    _identify_reseller_id,
    _prepare_file,
    _process_batch,
    apply_pricelist_to_marketplace,
    identify_marketplaces, validate_pricelist_batch,
)


@pytest.fixture
@patch.object(CBCService, 'parse_price_file')
@patch.object(CBCService, 'prepare_price_proposal')
@patch.object(CBCService, 'apply_prices')
@patch.object(CBCService, '__init__')
def cbc_service(
    mock___init__,
    mock_apply_prices,
    mock_prepare_price_proposal,
    mock_parse_price_file,
    parse_price_file_response,
    price_proposal_response,
):
    mock_parse_price_file.return_value = parse_price_file_response
    mock_prepare_price_proposal.return_value = price_proposal_response
    mock_apply_prices.return_value = 'Request Accepted.'
    mock___init__.return_value = None

    return CBCService()


@patch('connect_ext_ppr.services.pricing._fetch_batch_output_file')
def test_validate_pricelist_positive(mock_fetch_file):
    excel_file = NamedTemporaryFile(suffix='.xlsx')
    wb = openpyxl.load_workbook('./tests/fixtures/MFL-0000-0000-0000.xlsx')
    wb.save(excel_file)
    excel_file.seek(0)

    mock_fetch_file.return_value = (excel_file, 'filename.xlsx')

    validate_pricelist_batch(excel_file, 'BAT-1')

    assert excel_file.closed


@pytest.mark.parametrize('col_id,col_name', (
    (1, 'Billing Period'),
    (4, 'MPN'),
    (16, 'Effective Date'),
))
@patch('connect_ext_ppr.services.pricing._fetch_batch_output_file')
def test_validate_pricelist_negative_missed_req_cols(
    mock_fetch_file,
    col_id,
    col_name,
):
    excel_file = NamedTemporaryFile(suffix='.xlsx')
    wb = openpyxl.load_workbook('./tests/fixtures/MFL-0000-0000-0000.xlsx')
    wb['Data'].delete_cols(col_id)
    wb.save(excel_file)
    excel_file.seek(0)

    mock_fetch_file.return_value = (excel_file, 'filename.xlsx')

    with pytest.raises(ClientError) as e:
        validate_pricelist_batch(excel_file, 'BAT-1')

    assert excel_file.closed

    assert e.value.message == (
        "Pricing Batch output 'BAT-1' does not "
        f"contain mandatory column: {col_name}."
    )


@patch('connect_ext_ppr.services.pricing._fetch_batch_output_file')
def test_validate_pricelist_negative_no_cost_price(mock_fetch_file):
    excel_file = NamedTemporaryFile(suffix='.xlsx')
    wb = openpyxl.load_workbook('./tests/fixtures/MFL-0000-0000-0000.xlsx')
    wb['Data'].delete_cols(7)   # Cost
    wb['Data'].delete_cols(16)  # Price
    wb.save(excel_file)
    excel_file.seek(0)

    mock_fetch_file.return_value = (excel_file, 'filename.xlsx')

    with pytest.raises(ClientError) as e:
        validate_pricelist_batch(excel_file, 'BAT-1')

    assert e.value.message == (
        "Pricing Batch output 'BAT-1' does not "
        "contain either Cost or Price column."
    )


@pytest.mark.parametrize('col_id,col_name', (
    (12, 'Cost Currency'),
    (13, 'Price Currency'),
))
@patch('connect_ext_ppr.services.pricing._fetch_batch_output_file')
def test_validate_pricelist_negative_no_cost_price_currency(
    mock_fetch_file,
    col_id,
    col_name,
):
    excel_file = NamedTemporaryFile(suffix='.xlsx')
    wb = openpyxl.load_workbook('./tests/fixtures/MFL-0000-0000-0000.xlsx')
    wb['Data'].delete_cols(col_id)  # Cost
    wb.save(excel_file)
    excel_file.seek(0)

    mock_fetch_file.return_value = (excel_file, 'filename.xlsx')

    with pytest.raises(ClientError) as e:
        validate_pricelist_batch(excel_file, 'BAT-1')

    assert e.value.message == (
        "Pricing Batch output 'BAT-1' does not "
        f"contain mandatory column: {col_name}."
    )


@pytest.mark.parametrize('col_id,col_name,value', (
    (7, 'Cost', None),
    (17, 'Price', 'not number'),
    (4, 'MPN', None),
))
@patch('connect_ext_ppr.services.pricing._fetch_batch_output_file')
def test_validate_pricelist_negative_invalid_value(
    mock_fetch_file,
    col_id,
    col_name,
    value,
):
    excel_file = NamedTemporaryFile(suffix='.xlsx')
    wb = openpyxl.load_workbook('./tests/fixtures/MFL-0000-0000-0000.xlsx')
    wb['Data'].cell(row=2, column=col_id).value = value
    wb.save(excel_file)
    excel_file.seek(0)

    mock_fetch_file.return_value = (excel_file, 'filename.xlsx')

    with pytest.raises(ClientError) as e:
        validate_pricelist_batch(excel_file, 'BAT-1')

    assert e.value.message == (
        "Pricing Batch output 'BAT-1' contains invalid "
        f"value at column '{col_name}' of row '2'."
    )


@patch('connect_ext_ppr.services.pricing._fetch_batch_output_file')
def test_validate_pricelist_negative_invalid_effective_date(mock_fetch_file):
    excel_file = NamedTemporaryFile(suffix='.xlsx')
    wb = openpyxl.load_workbook('./tests/fixtures/MFL-0000-0000-0000.xlsx')
    wb['Data'].cell(row=2, column=16).value = 'Not a date'
    wb.save(excel_file)
    excel_file.seek(0)

    mock_fetch_file.return_value = (excel_file, 'filename.xlsx')

    with pytest.raises(ClientError) as e:
        validate_pricelist_batch(excel_file, 'BAT-1')

    assert e.value.message == (
        "Effective date 'Not a date' is either not "
        "found or invalid for first row in Batch BAT-1."
    )


def test_apply_pricelist_to_marketplace_positive(
    mocker,
    logger,
    marketplace,
    cbc_service,
    batch_output_file,
    connect_client,
    client_mocker_factory,
    deployment_request_factory,
    marketplace_config_factory,
):
    dep_req = deployment_request_factory()
    mp_conf = marketplace_config_factory('MP-123', deployment_request=dep_req)

    price_excel_file = NamedTemporaryFile(suffix='.xlsx')
    price_excel_file.read()
    price_excel_file.seek(0)

    reseller_id_mock = mocker.patch(
        'connect_ext_ppr.services.pricing._identify_reseller_id',
        return_value=marketplace['hubs'][0]['external_id'],
    )
    prepare_file_mock = mocker.patch(
        'connect_ext_ppr.services.pricing._prepare_file',
        return_value=(
            price_excel_file,
            'MFL-001.xlsx',
            {
                'cost': True,
                'price': True,
                'msrp': True,
                'effective_date': '07/26/2023',
            },
        ),
    )
    process_batch_mock = mocker.patch(
        'connect_ext_ppr.services.pricing._process_batch',
        return_value=None,
    )

    create_ppr_to_media_mock = mocker.patch(
        'connect_ext_ppr.services.pricing.create_ppr_to_media',
    )

    apply_pricelist_to_marketplace(
        dep_req,
        cbc_service,
        connect_client,
        mp_conf,
        logger,
    )

    reseller_id_mock.assert_called_once_with(
        client=connect_client,
        batch_id=mp_conf.pricelist_id,
        marketplace_id=mp_conf.marketplace,
        hub_id=dep_req.deployment.hub_id,
    )
    prepare_file_mock.assert_called_once_with(
        client=connect_client,
        batch_id=mp_conf.pricelist_id,
    )
    process_batch_mock.assert_called_once_with(
        cbc_service=cbc_service,
        excel_file=price_excel_file,
        file_name='MFL-001.xlsx',
        reseller_id=marketplace['hubs'][0]['external_id'],
        deployment=dep_req.deployment,
        dataset={
            'cost': True,
            'price': True,
            'msrp': True,
            'effective_date': '07/26/2023',
        },
        send_log=ANY,
    )

    assert price_excel_file

    create_ppr_to_media_mock.assert_called_once_with(
        client=ANY,
        account_id=dep_req.deployment.account_id,
        instance_id=dep_req.id,
        filename=f'{dep_req.id}_PriceApply_MP-123_None.xlsx',
        content=ANY,
    )


def test_identify_marketplaces_positive(
    connect_client,
    client_mocker_factory,
    marketplace,
):
    hub_id = 'HB-000-000'
    client_mocker = client_mocker_factory(base_url=connect_client.endpoint)
    client_mocker.marketplaces.filter(
        R().hubs.id.eq(hub_id),
    ).mock(
        return_value=[marketplace],
    )

    marketplace_ids = identify_marketplaces(connect_client, hub_id)

    TestCase().assertListEqual(marketplace_ids, [marketplace['id']])


def test_identify_marketplaces_negative_no_marketplace(
    connect_client,
    client_mocker_factory,
):
    hub_id = 'HB-000-000'

    client_mocker = client_mocker_factory(base_url=connect_client.endpoint)
    client_mocker.marketplaces.filter(
        R().hubs.id.eq(hub_id),
    ).mock(
        return_value=[],
    )

    with pytest.raises(ClientError):
        identify_marketplaces(connect_client, hub_id)


@responses.activate
def test_prepare_file_positive(
    connect_client,
    batch_file,
    batch_output_file,
):
    batch_id = 'BAT-0000-0000-0000'

    responses.add(
        method='GET',
        url=f'{connect_client.endpoint}/pricing/batches/BAT-0000-0000-0000/files?'
            f'eq(type,output)&limit=100&offset=0',
        json=[batch_file],
    )
    responses.add(
        method='GET',
        url=f'{connect_client.endpoint}/media/folders/streams_batches/BAT-0000-0000-0000/'
            f'files/MFL-0000-0000-0000/BAT-0000-0000-0000-out.xlsx',
        body=batch_output_file.read(),
        status=200,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    excel_file, file_name, dataset = _prepare_file(connect_client, batch_id)

    assert len(excel_file.read())
    assert file_name == f"{batch_file['id']}.xlsx"
    TestCase().assertDictEqual(
        dataset,
        {
            'cost': True,
            'price': True,
            'msrp': True,
            'effective_date': '07/26/2023',
        },
    )


def test_prepare_file_negative_no_batch_file(
    connect_client,
    client_mocker_factory,
):
    batch_id = 'BAT-0000-0000-0000'

    client_mocker = client_mocker_factory(base_url=connect_client.endpoint)
    client_mocker('pricing').batches[batch_id].files.filter(
        type='output',
    ).mock(
        return_value=[],
    )

    with pytest.raises(ClientError):
        _prepare_file(connect_client, batch_id)


def test_prepare_file_negative_more_than_one_batch_file(
    connect_client,
    client_mocker_factory,
    batch_file,
):
    batch_id = 'BAT-0000-0000-0000'

    client_mocker = client_mocker_factory(base_url=connect_client.endpoint)
    client_mocker('pricing').batches[batch_id].files.filter(
        type='output',
    ).mock(
        return_value=[batch_file, batch_file],
    )

    with pytest.raises(ClientError):
        _prepare_file(connect_client, batch_id)


def test_fetch_and_validate_batch_positive(
    connect_client,
    client_mocker_factory,
    batch,
    no_db_deployment,
):
    batch_id = 'BAT-0000-0000-0000'

    client_mocker = client_mocker_factory(base_url=connect_client.endpoint)
    client_mocker('pricing').batches.filter(
        id=batch_id,
    ).select(
        '+stream.context',
    ).mock(
        return_value=[batch],
    )

    response = _fetch_and_validate_batch(
        connect_client,
        batch_id,
        no_db_deployment,
    )

    TestCase().assertDictEqual(response, batch)


def test_fetch_and_validate_batch_negative_no_batch(
    connect_client,
    client_mocker_factory,
):
    batch_id = 'BAT-0000-0000-0000'

    client_mocker = client_mocker_factory(base_url=connect_client.endpoint)
    client_mocker('pricing').batches.filter(
        id=batch_id,
    ).select(
        '+stream.context',
    ).mock(
        return_value=[],
    )

    with pytest.raises(ClientError):
        _fetch_and_validate_batch(
            connect_client,
            batch_id,
            None,
        )


def test_fetch_and_validate_batch_negative_batch_does_not_belong_to_deployment(
    connect_client,
    client_mocker_factory,
    batch,
    no_db_deployment,
):
    batch_id = 'BAT-0000-0000-0000'
    deployment = deepcopy(no_db_deployment)
    deployment.product_id = 'PRD-NO-MATCH'

    client_mocker = client_mocker_factory(base_url=connect_client.endpoint)
    client_mocker('pricing').batches.filter(
        id=batch_id,
    ).select(
        '+stream.context',
    ).mock(
        return_value=[batch],
    )

    with pytest.raises(ClientError):
        _fetch_and_validate_batch(
            connect_client,
            batch_id,
            deployment,
        )


def test_identify_reseller_id_positive(
    connect_client,
    client_mocker_factory,
    batch,
    no_db_deployment,
    marketplace,
    hub,
):
    marketplace_id = batch['stream']['context']['marketplace']['id']

    client_mocker = client_mocker_factory(base_url=connect_client.endpoint)
    client_mocker.marketplaces[marketplace_id].get(
        return_value=marketplace,
    )
    client_mocker.hubs.filter(
        R().id.in_([hub['hub']['id'] for hub in marketplace['hubs']]),
        R().instance.type.eq('OA'),
    ).mock(
        return_value=[hub],
    )

    reseller_id = _identify_reseller_id(
        connect_client,
        batch['id'],
        marketplace_id,
        no_db_deployment.hub_id,
    )

    assert reseller_id == marketplace['hubs'][0]['external_id']


def test_identify_reseller_id_negative_deployment_hub_not_present_in_batch_marketplace(
    connect_client,
    client_mocker_factory,
    batch,
    no_db_deployment,
    marketplace,
):
    marketplace_id = batch['stream']['context']['marketplace']['id']

    client_mocker = client_mocker_factory(base_url=connect_client.endpoint)
    client_mocker.marketplaces[marketplace_id].get(
        return_value=marketplace,
    )
    client_mocker.hubs.filter(
        R().id.in_([hub['hub']['id'] for hub in marketplace['hubs']]),
        R().instance.type.eq('OA'),
    ).mock(
        return_value=[],
    )

    with pytest.raises(ClientError):
        _identify_reseller_id(
            connect_client,
            batch['id'],
            marketplace_id,
            no_db_deployment.hub_id,
        )


def test_identify_reseller_id_negative_hubs_not_configured_in_batch_marketplace(
    connect_client,
    client_mocker_factory,
    batch,
    no_db_deployment,
    marketplace,
):
    marketplace_id = batch['stream']['context']['marketplace']['id']
    modified_marketplace = deepcopy(marketplace)
    modified_marketplace.pop('hubs')

    client_mocker = client_mocker_factory(base_url=connect_client.endpoint)
    client_mocker.marketplaces[marketplace_id].get(
        return_value=modified_marketplace,
    )

    with pytest.raises(ClientError):
        _identify_reseller_id(
            connect_client,
            batch['id'],
            marketplace_id,
            no_db_deployment.hub_id,
        )


def test_get_reseller_id_negative_hubs_not_configured(
    marketplace,
):
    modified_marketplace = deepcopy(marketplace)
    modified_marketplace.pop('hubs')

    reseller_id = _get_reseller_id(modified_marketplace, 'HUB-0000-0000')

    assert reseller_id is None


def test_identify_reseller_id_negative_reseller_id_not_mapped(
    connect_client,
    client_mocker_factory,
    batch,
    no_db_deployment,
    marketplace,
    hub,
):
    marketplace_id = batch['stream']['context']['marketplace']['id']
    modified_marketplace = deepcopy(marketplace)
    modified_marketplace['hubs'][0].pop('external_id')

    client_mocker = client_mocker_factory(base_url=connect_client.endpoint)
    client_mocker.marketplaces[marketplace_id].get(
        return_value=modified_marketplace,
    )
    client_mocker.hubs.filter(
        R().id.in_([hub['hub']['id'] for hub in marketplace['hubs']]),
        R().instance.type.eq('OA'),
    ).mock(
        return_value=[hub],
    )

    with pytest.raises(ClientError):
        _identify_reseller_id(
            connect_client,
            batch['id'],
            marketplace_id,
            no_db_deployment.hub_id,
        )


@patch.object(CBCService, 'parse_price_file')
@patch.object(CBCService, 'prepare_price_proposal')
@patch.object(CBCService, 'apply_prices')
@patch.object(CBCService, '__init__')
def test_process_batch_positive(
    mock___init__,
    mock_apply_prices,
    mock_prepare_price_proposal,
    mock_parse_price_file,
    parse_price_file_response,
    price_proposal_response,
    hub_credentials,
    no_db_deployment,
    batch_dataset,
    logger,
):
    reseller_id = '10000001'

    mock_parse_price_file.return_value = parse_price_file_response
    mock_prepare_price_proposal.return_value = price_proposal_response
    mock_apply_prices.return_value = 'Request Accepted.'
    mock___init__.return_value = None

    excel_file = NamedTemporaryFile(suffix='.xlsx')
    excel_file.write(open('./tests/fixtures/MFL-0000-0000-0000.xlsx', 'rb').read())
    excel_file.seek(0)

    data_id = _process_batch(
        CBCService(hub_credentials),
        excel_file,
        'MFL-0000-0000-0000.xlsx',
        reseller_id,
        no_db_deployment,
        batch_dataset,
        logger,
    )

    assert data_id == parse_price_file_response['dataId']

    assert mock_parse_price_file.call_args[0] == (
        '10000001',
        'VA-000-000',
        excel_file,
    )
    assert mock_prepare_price_proposal.call_args[0] == (
        '10000001',
        {
            'status': 'PARSED',
            'priceListStructure': [
                'MPN', 'Vendor ID', 'Vendor Name', 'Service Template / Product Line',
                'Product', 'Billing Period', 'Subscription Period', 'Billing Model',
                'UOM', 'Lower Limit', 'Effective Date', 'Cost Billing Period',
                'Cost Currency', 'Cost', 'Price Currency', 'Price', 'MSRP',
                'Margin', 'Reseller Margin', 'Fee Type', 'Subscriptions',
                'Seats', 'Active',
            ],
            'pricingModel': 'FLAT',
            'feeType': 'RECURRING',
            'vendorId': 'VA-000-000',
            'dataId': 1,
        },
        True,
        True,
        True,
        '07/26/2023',
    )
    assert mock_apply_prices.call_args[0] == (
        '10000001',
        {
            'status': 'PARSED',
            'priceListStructure': [
                'MPN', 'Vendor ID', 'Vendor Name', 'Service Template / Product Line',
                'Product', 'Billing Period', 'Subscription Period',
                'Billing Model', 'UOM', 'Lower Limit', 'Effective Date',
                'Cost Billing Period', 'Cost Currency', 'Cost',
                'Price Currency', 'Price', 'MSRP', 'Margin',
                'Reseller Margin', 'Fee Type', 'Subscriptions',
                'Seats', 'Active',
            ],
            'pricingModel': 'FLAT',
            'feeType': 'RECURRING',
            'vendorId': 'VA-000-000',
            'dataId': 1,
        },
        True,
        True,
        True,
        '07/26/2023',
        'MFL-0000-0000-0000.xlsx',
    )

    assert logger.call_count == 3


def test_determine_dataset_negative_effective_date_column_not_present():
    wb = load_workbook(
        filename='./tests/fixtures/MFL-0000-0000-0001.xlsx',
    )
    ws = wb['Data']
    with pytest.raises(ClientError):
        _determine_dataset(ws, 'BAT-000-000-000')


def test_determine_dataset_negative_empty_effective_date():
    wb = load_workbook(
        filename='./tests/fixtures/MFL-0000-0000-0002.xlsx',
    )
    ws = wb['Data']
    with pytest.raises(ClientError):
        _determine_dataset(ws, 'BAT-000-000-000')


def test_determine_dataset_negative_effective_date_wrong_format():
    wb = load_workbook(
        filename='./tests/fixtures/MFL-0000-0000-0003.xlsx',
    )
    ws = wb['Data']
    with pytest.raises(ClientError):
        _determine_dataset(ws, 'BAT-000-000-000')
