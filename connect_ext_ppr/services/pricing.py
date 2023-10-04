from datetime import datetime
from tempfile import NamedTemporaryFile
from typing import Dict

from connect.client import ConnectClient
from connect.client.rql import R
from openpyxl import load_workbook

from connect_ext_ppr.errors import PriceUpdateError
from connect_ext_ppr.client.exception import CBCClientError
from connect_ext_ppr.utils import execute_with_retry, create_dr_file_to_media

PRICELIST_COLUMNS = {
    'MPN': str,
    'Billing Period': str,
    'Cost': (int, float),
    'Cost Currency': str,
    'Price': (int, float),
    'Price Currency': str,
    'MSRP': (int, float),
    'Effective Date': str,
}


def validate_pricelist_batch(connect_client, batch_id):
    """
    @param Client connect_client:
    @param str batch_id:

    @returns None
    @raises ClientError
    """
    excel_file, file_name = _fetch_batch_output_file(connect_client, batch_id)
    try:
        wb = load_workbook(excel_file)
        ws = wb['Data']
        _determine_dataset(ws, batch_id)
        wb.close()
    finally:
        excel_file.close()


def apply_pricelist_to_marketplace(
    deployment_request,
    cbc_service,
    connect_client,
    marketplace,
    logger,
):
    """
    @param DeploymentRequest deployment_request:
    @param CBCService cbc_service:
    @param Client connect_client:
    @param MarketplaceConfiguration marketplace:

    @returns None
    @raises ClientError, CBCClientError
    """
    logger.info(
        'Pricelist uploading %s_%s_%s: started.',
        deployment_request.id,
        marketplace.pricelist_id,
        marketplace.marketplace,
    )

    reseller_id = _identify_reseller_id(
        client=connect_client,
        batch_id=marketplace.pricelist_id,
        marketplace_id=marketplace.marketplace,
        hub_id=deployment_request.deployment.hub_id,
    )
    logger.info(
        'Pricelist uploading %s_%s_%s: reseller id %s.',
        deployment_request.id,
        marketplace.pricelist_id,
        marketplace.marketplace,
        reseller_id,
    )
    excel_file, file_name, dataset = _prepare_file(
        client=connect_client,
        batch_id=marketplace.pricelist_id,
    )
    create_dr_file_to_media(
        client=connect_client,
        account_id=deployment_request.deployment.account_id,
        dr_id=deployment_request.id,
        filename=file_name,
        content=excel_file,
    )
    excel_file.seek(0)

    logger.info(
        'Pricelist uploading %s_%s_%s: filename "%s", dataset "%s".',
        deployment_request.id,
        marketplace.pricelist_id,
        marketplace.marketplace,
        file_name,
        dataset,
    )
    try:
        data_id = _process_batch(
            cbc_service=cbc_service,
            excel_file=excel_file,
            file_name=file_name,
            reseller_id=reseller_id,
            deployment=deployment_request.deployment,
            dataset=dataset,
            logger=logger,
        )
        logger.info(
            'Pricelist uploading %s_%s_%s: data_id "%s".',
            deployment_request.id,
            marketplace.pricelist_id,
            marketplace.marketplace,
            data_id,
        )
    finally:
        excel_file.close()


def identify_marketplaces(
    client: ConnectClient,
    hub_id: str,
):
    marketplaces = client.marketplaces.filter(
        R().hubs.id.eq(hub_id),
    )

    if not marketplaces:
        raise PriceUpdateError.PLT_008(
            format_kwargs={'hub_id': hub_id},
        )

    return [marketplace['id'] for marketplace in marketplaces]


def _determine_dataset(ws, batch_id):
    dataset = {
        'cost': False,
        'price': False,
        'msrp': False,
        'effective_date': None,
    }

    headers = {
        cell.value: i
        for i, cell in enumerate(ws[1])
        if cell.value in PRICELIST_COLUMNS
    }

    dataset['msrp'] = 'MSRP' in headers

    _validate_required_columns(headers.keys(), batch_id)

    _validate_value_columns(headers.keys(), batch_id)

    dataset.update(_validate_value_columns(headers.keys(), batch_id))

    _validate_pricelist_content(ws, headers, batch_id)

    dataset['effective_date'] = _parse_effective_date(ws, headers, batch_id)

    return dataset


def _validate_required_columns(columns, batch_id):
    for required_col in ['MPN', 'Billing Period', 'Effective Date']:
        if required_col not in columns:
            raise PriceUpdateError.PLT_010(format_kwargs={
                'batch_id': batch_id,
                'col_name': required_col,
            })


def _validate_value_columns(columns, batch_id):
    dataset = {
        'cost': False,
        'price': False,
    }

    for val_col in ['Cost', 'Price']:
        if val_col in columns:
            dataset[val_col.lower()] = True
            if f'{val_col} Currency' not in columns:
                raise PriceUpdateError.PLT_010(format_kwargs={
                    'batch_id': batch_id,
                    'col_name': f'{val_col} Currency',
                })

    if (not dataset['cost']) and (not dataset['price']):
        raise PriceUpdateError.PLT_005(format_kwargs={
            'batch_id': batch_id,
        })

    return dataset


def _validate_pricelist_content(ws, columns, batch_id):
    for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for col in columns:
            if not isinstance(row[columns[col]].value, PRICELIST_COLUMNS[col]):
                raise PriceUpdateError.PLT_011(format_kwargs={
                    'batch_id': batch_id,
                    'column': col,
                    'row': row_number,
                })


def _parse_effective_date(ws, columns, batch_id):
    effective_date_str = ws[2][columns['Effective Date']].value

    try:
        effective_date = effective_date_str.split('T')[0]
        return datetime.strptime(
            effective_date,
            '%Y-%m-%d',
        ).strftime('%m/%d/%Y')
    except (ValueError, KeyError):
        raise PriceUpdateError.PLT_006(
            format_kwargs={
                'batch_id': batch_id,
                'date': effective_date_str,
            },
        )


def _identify_cbc_hubs(
    client: ConnectClient,
    marketplace: Dict,
):
    if 'hubs' in marketplace.keys():
        hub_ids = [hub['hub']['id'] for hub in marketplace['hubs']]

        osa_hubs = client.hubs.filter(
            R().id.in_(hub_ids),
            R().instance.type.eq('OA'),
        )

        return osa_hubs

    return []


def _get_reseller_id(marketplace, hub_id):
    hub_details = None
    if 'hubs' in marketplace.keys():
        hub_details = next(filter(lambda h: h['hub']['id'] == hub_id, marketplace['hubs']))

    if hub_details and 'external_id' in hub_details.keys():
        return hub_details['external_id']


def _prepare_file(client, batch_id):
    excel_file, file_name = _fetch_batch_output_file(client, batch_id)
    wb = load_workbook(excel_file)
    try:
        ws = wb['Data']
        dataset = _determine_dataset(ws, batch_id)
        ws.title = 'Price-list'
        excel_file.seek(0)
        wb.save(excel_file)
        excel_file.seek(0)
    finally:
        wb.close()

    return excel_file, file_name, dataset


# TODO: no used
def _fetch_and_validate_batch(client, batch_id, deployment):
    batches = client('pricing').batches.filter(id=batch_id).select('+stream.context')
    if not batches:
        raise PriceUpdateError.PLT_001(
            format_kwargs={'batch_id': batch_id},
        )

    batch = batches[0]

    product_id = batch['stream']['context']['product']['id']

    if product_id != deployment.product_id:
        raise PriceUpdateError.PLT_009(
            format_kwargs={
                'batch_id': batch_id,
                'b_product_id': product_id,
                'deployment_id': deployment.id,
                'd_product_id': deployment.product_id,
            },
        )

    return batch


def _fetch_batch_output_file(client, batch_id):
    files = list(client('pricing').batches[batch_id].files.filter(type='output'))
    if not files:
        raise PriceUpdateError.PLT_007(
            format_kwargs={'batch_id': batch_id},
        )

    if len(files) > 1:
        raise PriceUpdateError.PLT_002(
            format_kwargs={'batch_id': batch_id},
        )

    file = files[0]
    file_location = file['name']
    file_content = client.get(
        file_location[11:] if file_location.startswith('/public/v1/') else file_location,
    )
    file_name = f"{file['id']}.xlsx"

    excel_file = NamedTemporaryFile(suffix='.xlsx')
    excel_file.write(file_content)
    excel_file.seek(0)

    return excel_file, file_name


def _identify_reseller_id(client, batch_id, marketplace_id, hub_id):
    marketplace = client.marketplaces[marketplace_id].get()

    marketplace_hubs = _identify_cbc_hubs(client, marketplace)

    marketplace_hub_ids = [hub['id'] for hub in marketplace_hubs]

    if hub_id not in marketplace_hub_ids:
        raise PriceUpdateError.PLT_003(
            format_kwargs={
                'hub_id': hub_id,
                'marketplace_id': marketplace_id,
                'batch_id': batch_id,
            },
        )

    reseller_id = _get_reseller_id(marketplace, hub_id)

    if not reseller_id:
        raise PriceUpdateError.PLT_004(
            format_kwargs={
                'hub_id': hub_id,
                'marketplace_id': marketplace_id,
            },
        )

    return reseller_id


def _process_batch(
    cbc_service,
    excel_file,
    file_name,
    reseller_id,
    deployment,
    dataset,
    logger,
):
    excel_file.seek(0)

    parsed_price = execute_with_retry(
        function=cbc_service.parse_price_file,
        exception_class=CBCClientError,
        args=(reseller_id, deployment.vendor_id, excel_file),
    )
    logger.info('Parsed price: %s', parsed_price)

    data_id = parsed_price['dataId']

    execute_with_retry(
        function=cbc_service.prepare_price_proposal,
        exception_class=CBCClientError,
        args=(
            reseller_id, parsed_price, dataset['cost'],
            dataset['price'], dataset['msrp'], dataset['effective_date'],
        ),
    )

    execute_with_retry(
        function=cbc_service.apply_prices,
        exception_class=CBCClientError,
        args=(
            reseller_id,
            parsed_price,
            dataset['cost'],
            dataset['price'],
            dataset['msrp'],
            dataset['effective_date'],
            file_name,
        ),
    )

    return data_id
